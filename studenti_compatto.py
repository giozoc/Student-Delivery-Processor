#!/usr/bin/env python3
from __future__ import annotations

import io
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import xlrd
import pymupdf as fitz
from PIL import Image, ImageOps
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# File Excel da leggere nella cartella root
EXCEL_FILES = ["AA2024-2025.xls", "AA2025-2026.xls"]

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
PDF_EXTS = {".pdf"}


@dataclass
class Record:
    cognome_nome: str
    anno_accademico: str
    matricola: str


def norm_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", str(value or "").strip())
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.replace("’", "'").replace("`", "'")
    value = re.sub(r"[_\-\s]+", " ", value).strip()
    return value.upper()


def norm_token(token: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", norm_text(token).replace("'", ""))


def token_counter(value: str) -> Counter:
    return Counter(t for t in map(norm_token, norm_text(value).split()) if t)


def norm_matricola(value: str) -> str:
    value = str(value or "").strip()
    return re.sub(r"\D+", "", value[:-2] if value.endswith(".0") else value)


def before_underscore(name: str) -> str:
    return str(name or "").split("_", 1)[0].strip()


def root_dir(root: str | Path = ".") -> Path:
    root = Path(root).expanduser()
    if str(root).strip() in {"", "."}:
        return Path(__file__).parent
    return root if root.is_absolute() else Path(__file__).parent / root


def year_from_xls(path: Path) -> str:
    sh = xlrd.open_workbook(str(path)).sheet_by_index(0)
    for r in range(min(sh.nrows, 20)):
        if str(sh.cell_value(r, 0)).strip().upper() == "ANNO ACCADEMICO":
            return str(sh.cell_value(r, 1)).strip()
    return ""


def records_from_xls(path: Path) -> list[Record]:
    sh = xlrd.open_workbook(str(path)).sheet_by_index(0)
    anno = year_from_xls(path)

    header_row, header = None, {}
    for r in range(min(sh.nrows, 50)):
        row = [str(sh.cell_value(r, c)).strip().upper() for c in range(sh.ncols)]
        if "COGNOME E NOME" in row and "MATRICOLA" in row:
            header_row = r
            header = {name: i for i, name in enumerate(row)}
            break
    if header_row is None:
        raise RuntimeError(f"Intestazione non trovata in {path.name}")

    col_name, col_mat = header["COGNOME E NOME"], header["MATRICOLA"]
    out = []
    for r in range(header_row + 1, sh.nrows):
        name = re.sub(r"\s+", " ", str(sh.cell_value(r, col_name)).strip())
        if not name or norm_text(name) == "TOTALE":
            continue
        out.append(Record(name, anno, norm_matricola(sh.cell_value(r, col_mat))))
    return out


def build_indexes(xls_files: list[Path]):
    by_matricola, by_name, all_records = {}, {}, []
    for xls in xls_files:
        for rec in records_from_xls(xls):
            all_records.append(rec)
            if rec.matricola:
                by_matricola.setdefault(rec.matricola, []).append(rec)
            by_name.setdefault(norm_text(rec.cognome_nome), []).append(rec)
    return by_matricola, by_name, all_records


def folder_numeric_id(folder_name: str) -> str:
    m = re.search(r"_(\d{4,12})(?:_|$)", folder_name)
    return m.group(1) if m else ""


def resolve_record(folder_name: str, by_matricola, by_name, all_records) -> Record | None:
    folder_id = folder_numeric_id(folder_name)
    if folder_id in by_matricola and len(by_matricola[folder_id]) == 1:
        return by_matricola[folder_id][0]

    plain = before_underscore(folder_name)
    exact = by_name.get(norm_text(plain), [])
    if len(exact) == 1:
        return exact[0]

    folder_tokens = token_counter(plain)
    matches = [r for r in all_records if token_counter(r.cognome_nome) == folder_tokens]
    if len(matches) == 1:
        return matches[0]

    folder_set = set(folder_tokens.elements())
    weak = [r for r in all_records if folder_set.issubset(set(token_counter(r.cognome_nome).elements()))]
    return weak[0] if len(weak) == 1 else None


def rename_folder(folder: Path, rec: Record) -> Path:
    target = folder.with_name(rec.cognome_nome)
    if target == folder or target.exists():
        return folder
    folder.rename(target)
    print(f"Rinominata cartella: '{folder.name}' -> '{target.name}'")
    return target


def pdf_output(folder: Path) -> Path:
    return folder / f"{folder.name.replace(' ', '_')}.pdf"


def resize_img(img: Image.Image, max_size: int) -> Image.Image:
    img = ImageOps.exif_transpose(img)
    if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
        base = Image.new("RGB", img.size, "white")
        rgba = img.convert("RGBA")
        base.paste(rgba, mask=rgba.getchannel("A"))
        img = base
    elif img.mode != "RGB":
        img = img.convert("RGB")
    if max(img.size) <= max_size:
        return img
    ratio = max_size / max(img.size)
    size = (max(1, int(img.width * ratio)), max(1, int(img.height * ratio)))
    return img.resize(size, Image.Resampling.LANCZOS)


def image_to_page(path: Path, max_size: int) -> Image.Image:
    with Image.open(path) as img:
        return resize_img(img, max_size).copy()


def pdf_to_pages(path: Path, dpi: int, quality: int, max_size: int) -> list[Image.Image]:
    pages = []
    doc = fitz.open(path)
    try:
        for page in doc:
            pix = page.get_pixmap(dpi=dpi, alpha=False)
            img = resize_img(Image.frombytes("RGB", [pix.width, pix.height], pix.samples), max_size)
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality, optimize=True)
            buf.seek(0)
            pages.append(Image.open(buf).convert("RGB").copy())
            buf.close()
    finally:
        doc.close()
    return pages


def supported_files(folder: Path, out_pdf: Path) -> list[Path]:
    return [
        p for p in sorted(folder.iterdir(), key=lambda x: x.name.lower())
        if p.is_file() and p.suffix.lower() in (IMAGE_EXTS | PDF_EXTS) and p.absolute() != out_pdf.absolute()
    ]


def build_pdf(folder: Path, dpi: int = 140, quality: int = 55, max_size: int = 1800, keep_originals: bool = False) -> bool:
    out_pdf = pdf_output(folder)
    files = supported_files(folder, out_pdf)
    if not files:
        print(f"Nessun file supportato in: {folder.name}")
        return False

    pages = []
    for f in files:
        pages.extend(pdf_to_pages(f, dpi, quality, max_size) if f.suffix.lower() in PDF_EXTS else [image_to_page(f, max_size)])
    if not pages:
        return False

    first, rest = pages[0], pages[1:]
    try:
        first.save(out_pdf, format="PDF", save_all=True, append_images=rest, resolution=150)
        print(f"Creato PDF: {out_pdf.name}")
    finally:
        for img in pages:
            img.close()

    if not keep_originals:
        for f in files:
            try:
                f.unlink()
            except Exception:
                pass
    return True


def write_output_excel(root: Path, rows: list[list[str]], filename: str = "studenti_compilati.xlsx") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Risultati"
    ws.append(["Cognome e nome", "Anno Accademico", "Matricola"])
    for row in rows:
        ws.append(row)

    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 40)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width

    out = root / filename
    wb.save(out)
    return out


def process_root(root: str | Path = ".", keep_originals: bool = False, output_excel: str = "studenti_compilati.xlsx") -> Path:
    root = root_dir(root)
    if not root.exists() or not root.is_dir():
        raise SystemExit(f"Cartella non valida: {root}")

    xls_files = [root / name for name in EXCEL_FILES]
    missing = [p.name for p in xls_files if not p.exists()]
    if missing:
        raise SystemExit("File Excel mancanti nella cartella root: " + ", ".join(missing))

    by_matricola, by_name, all_records = build_indexes(xls_files)
    rows = []

    for folder in sorted(root.iterdir(), key=lambda x: x.name.lower()):
        if not folder.is_dir() or folder.name.startswith(".") or folder.name == ".venv":
            continue

        rec = resolve_record(folder.name, by_matricola, by_name, all_records)
        if not rec:
            print(f"Cartella non trovata negli Excel: {folder.name}")
            continue

        folder = rename_folder(folder, rec)
        build_pdf(folder, keep_originals=keep_originals)
        rows.append([rec.cognome_nome, rec.anno_accademico, rec.matricola])

    out = write_output_excel(root, rows, filename=output_excel)
    print(f"Creato file Excel finale: {out}")
    return out


if __name__ == "__main__":
    process_root(keep_originals=False)
